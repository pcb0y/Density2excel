from openpyxl import load_workbook

# 打开Excel文件
wb = load_workbook('density_data.xlsx')

# 获取活动工作表
ws = wb.active
print('工作表名称:', ws.title)

# 查看表头
print('\n表头:')
for cell in ws[1]:
    print(cell.value, end='\t')

# 查看数据行
print('\n\n数据行:')
for row in ws.iter_rows(min_row=2, values_only=True):
    print(row)

# 关闭文件
wb.close()