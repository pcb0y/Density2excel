import serial
import csv
import re
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
from tkinter import filedialog
import threading
import configparser


def read_serial_data(port, baudrate=9600, bytesize=8, stopbits=serial.STOPBITS_ONE, parity=serial.PARITY_NONE, timeout=3):
    """
    从COM口读取数据
    :param port: 串口名称，如COM3（Windows）或/dev/ttyUSB0（Linux）
    :param baudrate: 波特率
    :param bytesize: 数据位
    :param stopbits: 停止位
    :param parity: 校验位
    :param timeout: 超时时间
    :return: 读取到的串口数据字符串
    """
    # 转换停止位
    if stopbits == 1:
        stopbits = serial.STOPBITS_ONE
    elif stopbits == 1.5:
        stopbits = serial.STOPBITS_ONE_POINT_FIVE
    elif stopbits == 2:
        stopbits = serial.STOPBITS_TWO
    
    # 转换校验位
    if parity == 'NONE':
        parity = serial.PARITY_NONE
    elif parity == 'ODD':
        parity = serial.PARITY_ODD
    elif parity == 'EVEN':
        parity = serial.PARITY_EVEN
    try:
        # 初始化串口，增加流控制设置
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            parity=parity,
            stopbits=stopbits,
            bytesize=bytesize,
            timeout=timeout,
            xonxoff=False,  # 禁用软件流控制
            rtscts=False,   # 禁用硬件流控制
            dsrdtr=False,   # 禁用DSR/DTR流控制
            writeTimeout=2
        )

        # 清空输入缓冲区，确保读取最新数据
        ser.flushInput()
        
        # 读取串口数据
        data = ""
        lines_read = 0
        max_lines = 10  # 增加最大读取行数，提高兼容性
        line_timeout = 0.5  # 每行读取的超时时间
        
        # 尝试读取max_lines行数据，或直到超时
        start_time = time.time()
        while time.time() - start_time < timeout and lines_read < max_lines:
            line_start_time = time.time()
            line_data = b""
            
            # 读取一行数据，处理超时
            while time.time() - line_start_time < line_timeout:
                if ser.in_waiting > 0:
                    byte = ser.read(1)
                    if byte == b'\n':
                        break
                    line_data += byte
                else:
                    time.sleep(0.01)  # 短暂休眠，减少CPU占用
            
            # 解码并处理读取到的行
            if line_data:
                line = line_data.decode('utf-8', errors='ignore').strip()
                if line:
                    data += line + "\n"
                    lines_read += 1
                    # 如果已经找到密度数据，可以提前返回
                    if "Density" in line:
                        # 再读取1-2行，确保获取完整数据
                        for _ in range(2):
                            if ser.in_waiting > 0:
                                extra_line = ser.readline().decode('utf-8', errors='ignore').strip()
                                if extra_line:
                                    data += extra_line + "\n"
                    
        # 关闭串口前检查是否还有剩余数据
        if ser.in_waiting > 0:
            remaining_data = ser.read(ser.in_waiting).decode('utf-8', errors='ignore').strip()
            if remaining_data:
                data += remaining_data + "\n"
        
        ser.close()
        
        # 改进数据完整性检查
        if data.strip():
            # 打印读取到的原始数据，用于调试
            # print(f"读取到的数据: {data}")
            return data
        else:
            return ""

    except Exception as e:
        # 不打印每次读取错误，避免控制台信息过多
        # print(f"串口读取错误: {e}")
        return ""


def extract_density_value(data):
    """
    从串口数据中提取密度值（如1.329）
    :param data: 串口读取的原始数据字符串
    :return: 提取到的密度值（浮点数），提取失败返回None
    """
    # 使用更灵活的正则表达式匹配密度值
    # 匹配 "Density" 或 "density" 后跟冒号和数值
    pattern = r'[Dd]ensity\s*:\s*(\d+\.\d+)\s*'
    match = re.search(pattern, data)

    if match:
        try:
            density_value = float(match.group(1))
            return density_value
        except ValueError:
            print("密度值转换为浮点数失败")
            return None
    else:
        # 如果未找到Density关键字，尝试直接提取所有浮点数
        number_pattern = r'(\d+\.\d+)'
        numbers = re.findall(number_pattern, data)
        if numbers:
            try:
                # 返回第一个匹配的浮点数
                return float(numbers[0])
            except ValueError:
                print("密度值转换为浮点数失败")
                return None
        print("未找到密度值")
        return None


def write_to_excel(data, filename="density_data.xlsx"):
    """
    将密度测试数据写入Excel文件
    :param data: 包含所有测试信息的字典
    :param filename: Excel文件名
    """
    try:
        # 检查文件是否存在
        file_exists = os.path.exists(filename)
        
        if file_exists:
            # 加载现有文件
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            # 创建新文件和工作表
            workbook = Workbook()
            sheet = workbook.active
            # 设置表头
            headers = ["来样时间", "测试时间", "机台号", "产品型号", "班次", 
                      "密度1", "密度2", "密度3", "密度4", "密度5", "平均值"]
            sheet.append(headers)
        
        # 准备要写入的数据行
        row_data = [
            data.get("来样时间", ""),
            data.get("测试时间", ""),
            data.get("机台号", ""),
            data.get("产品型号", ""),
            data.get("班次", ""),
            data.get("密度1", ""),
            data.get("密度2", ""),
            data.get("密度3", ""),
            data.get("密度4", ""),
            data.get("密度5", ""),
            data.get("平均值", "")
        ]
        
        # 写入数据行
        sheet.append(row_data)
        
        # 保存文件
        workbook.save(filename)
        print(f"成功将测试数据写入Excel文件: {filename}")

    except Exception as e:
        print(f"写入Excel文件错误: {e}")
        # 打印更详细的错误信息
        import traceback
        traceback.print_exc()


def read_product_models_from_excel(filename="density_data.xlsx"):
    """
    从Excel文件中读取产品型号、机台号等信息
    :param filename: Excel文件名
    :return: 产品型号列表
    """
    try:
        # 加载Excel文件
        workbook = load_workbook(filename)
        sheet = workbook.active
        
        product_info_list = []
        
        # 遍历所有行，从第2行开始（跳过表头）
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 检查产品型号是否存在（第4列）
            if row[3] and str(row[3]).strip():
                product_info = {
                    "来样时间": row[0] if row[0] else "",
                    "机台号": row[2] if row[2] else "",
                    "产品型号": str(row[3]).strip(),
                    "班次": row[4] if row[4] else ""
                }
                product_info_list.append(product_info)
        
        workbook.close()
        return product_info_list
        
    except Exception as e:
        print(f"从Excel读取产品型号错误: {e}")
        import traceback
        traceback.print_exc()
        return []


def update_excel_with_detection_results(filename, product_model, detect_data):
    """
    更新Excel文件中的检测结果
    :param filename: Excel文件名
    :param product_model: 产品型号
    :param detect_data: 检测数据字典
    """
    workbook = None
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active

        target_product = str(product_model).strip() if product_model is not None else ""
        detection_time = detect_data.get("检测时间")
        if detection_time is None:
            detection_time = detect_data.get("测试时间")

        updated = False
        for row in sheet.iter_rows(min_row=2):
            cell_value = row[3].value
            current_product = str(cell_value).strip() if cell_value is not None else ""
            if current_product == target_product and current_product:
                row[1].value = detection_time
                row[5].value = detect_data.get("密度1")
                row[6].value = detect_data.get("密度2")
                row[7].value = detect_data.get("密度3")
                row[8].value = detect_data.get("密度4")
                row[9].value = detect_data.get("密度5")
                row[10].value = detect_data.get("平均值")
                updated = True
                break

        if not updated and target_product:
            sheet.append([
                detect_data.get("来样时间", ""),
                detection_time,
                detect_data.get("机台号", ""),
                target_product,
                detect_data.get("班次", ""),
                detect_data.get("密度1"),
                detect_data.get("密度2"),
                detect_data.get("密度3"),
                detect_data.get("密度4"),
                detect_data.get("密度5"),
                detect_data.get("平均值"),
            ])

        workbook.save(filename)

    except Exception as e:
        print(f"更新Excel文件错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass


def update_excel_with_test_results(filename, product_model, test_data):
    detection_data = dict(test_data) if test_data is not None else {}
    if "检测时间" not in detection_data and "测试时间" in detection_data:
        detection_data["检测时间"] = detection_data.get("测试时间")
    return update_excel_with_detection_results(filename, product_model, detection_data)


def main():
    # 读取配置文件
    config = configparser.ConfigParser()
    config_file = "config.ini"
    
    serial_port = "COM2"
    baudrate = 9600
    bytesize = 7
    stopbits = 1
    parity = 'NONE'
    timeout = 2

    if os.path.exists(config_file):
        config.read(config_file)
        if config.has_section("SerialConfig"):
            serial_port = config["SerialConfig"].get("port", serial_port)
            baudrate = int(config["SerialConfig"].get("baudrate", str(baudrate)))
            bytesize = int(config["SerialConfig"].get("bytesize", str(bytesize)))
            stopbits = float(config["SerialConfig"].get("stopbits", str(stopbits)))
            parity = config["SerialConfig"].get("parity", parity)
            timeout = float(config["SerialConfig"].get("timeout", str(timeout)))
    
    excel_filename = "density_data.xlsx"
    
    print("密度检测系统启动")
    
    try:
        # 从Excel中读取所有产品型号
        product_info_list = read_product_models_from_excel(excel_filename)
        
        if not product_info_list:
            print("未从Excel文件中读取到产品型号，程序结束")
            return
        
        print(f"\n从Excel文件中读取到 {len(product_info_list)} 个产品型号:")
        for info in product_info_list:
            print(f"- {info['产品型号']} (机台号: {info['机台号']})")
        
        # 按顺序处理每个产品型号
        for i, product_info in enumerate(product_info_list, 1):
            product_model = product_info["产品型号"]
            machine_id = product_info["机台号"]
            sample_time = product_info["来样时间"]
            shift = product_info["班次"]
            
            print(f"\n=== 开始处理第 {i}/{len(product_info_list)} 个产品: {product_model} ===")
            print(f"请放入 {product_model} 型号的样块...")
            
            # 等待用户准备好
            input("准备就绪后按回车开始测试...")
            
            # 开始5次密度测试
            density_values = []
            test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for test_num in range(1, 6):
                print(f"\n开始第 {test_num} 次测试...")
                
                # 读取串口数据，最多尝试10次
                max_attempts = 10
                density = None
                
                for attempt in range(max_attempts):
                    raw_data = read_serial_data(serial_port, baudrate=baudrate, bytesize=bytesize, stopbits=stopbits, parity=parity, timeout=timeout)
                    if raw_data:
                        print("读取到的原始数据:")
                        print(raw_data)
                        
                        # 提取密度值
                        density = extract_density_value(raw_data)
                        if density is not None:
                            print(f"第 {test_num} 次测试成功提取密度值: {density} g/ccm")
                            break
                    else:
                        print(f"第 {attempt + 1} 次尝试读取失败，重试中...")
                        time.sleep(1)
                
                if density is not None:
                    density_values.append(density)
                else:
                    print(f"第 {test_num} 次测试失败，将使用None值")
                    density_values.append(None)
                
                # 等待用户准备下一次测试
                if test_num < 5:
                    input(f"第 {test_num} 次测试完成，请准备下一次测试，按回车继续...")
            
            # 计算平均值（仅包含有效数值）
            valid_densities = [d for d in density_values if d is not None]
            average_density = sum(valid_densities) / len(valid_densities) if valid_densities else None
            
            # 准备测试数据
            test_data = {
                "来样时间": sample_time,
                "检测时间": test_time,
                "机台号": machine_id,
                "产品型号": product_model,
                "班次": shift,
                "密度1": density_values[0] if len(density_values) > 0 else None,
                "密度2": density_values[1] if len(density_values) > 1 else None,
                "密度3": density_values[2] if len(density_values) > 2 else None,
                "密度4": density_values[3] if len(density_values) > 3 else None,
                "密度5": density_values[4] if len(density_values) > 4 else None,
                "平均值": round(average_density, 4) if average_density is not None else None
            }
            
            # 更新Excel文件
            update_excel_with_test_results(excel_filename, product_model, test_data)
            
            # 显示测试结果
            print("\n=== 测试结果 ===")
            print(f"产品型号: {product_model}")
            for j, d in enumerate(density_values, 1):
                print(f"密度{j}: {d} g/ccm" if d is not None else f"密度{j}: 测试失败")
            print(f"平均值: {test_data['平均值']} g/ccm" if test_data['平均值'] is not None else "平均值: 无法计算")
            print("===============")
        
        print("\n所有产品型号测试完成！")
        
    except KeyboardInterrupt:
        print("\n用户中断程序，退出测试系统")
    except Exception as e:
        print(f"程序运行出错: {e}")
        import traceback
        traceback.print_exc()


# 测试用：模拟完整的测试流程
def test_with_fixed_data():
    """模拟从Excel读取产品型号并进行测试的完整流程"""
    # 首先显示当前Excel文件内容
    print("=== 测试前Excel文件内容 ===")
    product_info_list = read_product_models_from_excel()
    for info in product_info_list:
        print(info)
    
    # 模拟测试数据
    test_data_str = """Air          :    +   7.5262 g
Liquid       :    +   1.8717 g
Volume       :         5.663 ccm
Density      :         1.329 g/ccm"""
    
    print("\n=== 开始模拟测试流程 ===")
    
    # 从Excel读取产品型号
    product_info_list = read_product_models_from_excel()
    
    if not product_info_list:
        print("未从Excel文件中读取到产品型号")
        return
    
    print(f"从Excel文件中读取到 {len(product_info_list)} 个产品型号:")
    for info in product_info_list:
        print(f"- {info['产品型号']} (机台号: {info['机台号']})")
    
    # 模拟处理前3个产品型号
    for i, product_info in enumerate(product_info_list[:3], 1):
        product_model = product_info["产品型号"]
        machine_id = product_info["机台号"]
        sample_time = product_info["来样时间"]
        shift = product_info["班次"]
        
        print(f"\n=== 模拟处理第 {i}/{len(product_info_list)} 个产品: {product_model} ===")
        print(f"请放入 {product_model} 型号的样块...")
        
        # 模拟用户准备就绪
        print("准备就绪后按回车开始测试... (模拟回车)")
        
        # 模拟5次密度测试
        density_values = []
        test_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for test_num in range(1, 6):
            print(f"\n开始第 {test_num} 次测试...")
            
            # 模拟读取串口数据
            print("读取到的原始数据:")
            print(test_data_str)
            
            # 提取密度值
            density = extract_density_value(test_data_str)
            if density is not None:
                print(f"第 {test_num} 次测试成功提取密度值: {density} g/ccm")
                # 添加一些随机波动使数据更真实
                import random
                density_with_variation = density + (random.random() - 0.5) * 0.05
                density_values.append(round(density_with_variation, 4))
            else:
                print(f"第 {test_num} 次测试失败")
                density_values.append(None)
            
            # 模拟用户准备下一次测试
            if test_num < 5:
                print(f"第 {test_num} 次测试完成，请准备下一次测试，按回车继续... (模拟回车)")
        
        # 计算平均值
        valid_densities = [d for d in density_values if d is not None]
        average_density = sum(valid_densities) / len(valid_densities) if valid_densities else None
        
        # 准备测试数据
        test_data = {
            "来样时间": sample_time,
            "测试时间": test_time,
            "机台号": machine_id,
            "产品型号": product_model,
            "班次": shift,
            "密度1": density_values[0],
            "密度2": density_values[1],
            "密度3": density_values[2],
            "密度4": density_values[3],
            "密度5": density_values[4],
            "平均值": round(average_density, 4) if average_density is not None else None
        }
        
        # 更新Excel文件
        update_excel_with_test_results("density_data.xlsx", product_model, test_data)
        
        # 显示测试结果
        print("\n=== 测试结果 ===")
        print(f"产品型号: {product_model}")
        for j, d in enumerate(density_values, 1):
            print(f"密度{j}: {d} g/ccm" if d is not None else f"密度{j}: 测试失败")
        print(f"平均值: {test_data['平均值']} g/ccm" if test_data['平均值'] is not None else "平均值: 无法计算")
        print("===============")
    
    # 显示测试后Excel文件内容
    print("\n=== 测试后Excel文件内容 ===")
    wb = load_workbook("density_data.xlsx")
    ws = wb.active
    print("表头:", [cell.value for cell in ws[1]])
    print("数据行:")
    for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
        print(row)
    wb.close()
    
    print("\n测试流程模拟完成！")

# 保留原有的CSV写入函数，便于向后兼容
def write_to_csv(value, filename="density_data.csv", header=["密度值(g/ccm)"]):
    """
    将密度值写入CSV文件（向后兼容）
    :param value: 要写入的密度值
    :param filename: CSV文件名
    :param header: CSV文件头
    """
    try:
        # 检查文件是否存在
        file_exists = os.path.exists(filename)
        
        # 写入数据
        with open(filename, 'a', newline='', encoding='utf-8') as f:
            # 使用明确的逗号分隔符
            writer = csv.writer(f, delimiter=',')
            # 如果文件是新的，先写入表头
            if not file_exists:
                writer.writerow(header)
            # 写入密度值
            writer.writerow([value])

        print(f"成功将密度值 {value} 写入CSV文件: {filename}")

    except Exception as e:
        print(f"写入CSV文件错误: {e}")
        # 打印更详细的错误信息
        import traceback
        traceback.print_exc()


class DensityDetectGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("密度检测系统")
        self.root.geometry("1000x700")
        self.root.resizable(True, True)
        
        # 设置mac风格主题
        self.style = ttk.Style()
        # 尝试使用clam主题，这是最接近mac风格的内置主题
        try:
            self.style.theme_use("clam")
        except:
            pass
        
        # Mac风格的颜色方案
        self.mac_colors = {
            "background": "#f2f2f7",
            "surface": "#ffffff",
            "text": "#000000",
            "primary": "#007aff",
            "secondary": "#8e8e93",
            "border": "#c6c6c8",
            "hover": "#0051d5",
            "active": "#e5e5ea"
        }
        
        # 自定义mac风格颜色和字体
        self.style.configure(
            "TFrame"
        )
        
        self.style.configure(
            "TLabel", 
            foreground=self.mac_colors["text"],
            font=("Segoe UI", 10)
        )
        
        self.style.configure(
            "TButton", 
            foreground=self.mac_colors["text"],
            font=("Segoe UI", 10),
            padding=8,
            relief="flat",
            borderwidth=0
        )
        
        self.style.map(
            "TButton", 
            foreground=[
                ("active", self.mac_colors["surface"]),
                ("!active", self.mac_colors["text"])
            ]
        )
        
        self.style.configure(
            "TEntry", 
            foreground=self.mac_colors["text"],
            font=("Segoe UI", 10),
            relief="flat",
            borderwidth=1,
            bordercolor=self.mac_colors["border"],
            lightcolor=self.mac_colors["primary"],
            darkcolor=self.mac_colors["primary"]
        )
        
        self.style.configure(
            "TCombobox", 
            foreground=self.mac_colors["text"],
            font=("Segoe UI", 10),
            relief="flat",
            borderwidth=1,
            bordercolor=self.mac_colors["border"]
        )
        
        self.style.configure(
            "TCombobox.Listbox",
            foreground=self.mac_colors["text"],
            font=("Segoe UI", 10),
            relief="flat"
        )
        
        self.style.map(
            "TCombobox",
            arrowcolor=[("active", self.mac_colors["primary"])]
        )
        
        self.style.configure(
            "TLabelFrame", 
            foreground=self.mac_colors["text"],
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            borderwidth=1,
            bordercolor=self.mac_colors["border"]
        )
        
        self.style.configure(
            "Treeview",
            font=("Segoe UI", 10)
        )
        
        self.style.configure(
            "Treeview.Heading",
            font=("Segoe UI", 10, "bold")
        )
        

        

        
        self.style.configure(
            "Vertical.TScrollbar",
            relief="flat",
            borderwidth=0
        )
        
        # 窗口背景色使用默认设置
        
        # 尝试设置窗口透明度（如果支持）
        try:
            self.root.attributes("-alpha", 0.98)
        except:
            pass
        
        # 读取配置文件
        self.config = configparser.ConfigParser()
        self.config_file = "config.ini"
        
        # 如果配置文件不存在，创建默认配置
        if not os.path.exists(self.config_file):
            self.config['SerialConfig'] = {
                'port': 'COM2',
                'baudrate': '9600',
                'bytesize': '7',
                'stopbits': '1',
                'parity': 'NONE',
                'timeout': '2',
                'max_attempts': '15'
            }
            with open(self.config_file, 'w') as f:
                self.config.write(f)
        else:
            # 读取配置文件
            self.config.read(self.config_file)
        
        # 设置全局变量
        self.serial_port = self.config['SerialConfig']['port']
        self.baudrate = int(self.config['SerialConfig']['baudrate'])
        self.bytesize = int(self.config['SerialConfig']['bytesize'])
        self.stopbits = float(self.config['SerialConfig']['stopbits'])
        self.parity = self.config['SerialConfig']['parity']
        self.timeout = float(self.config['SerialConfig']['timeout'])
        self.max_attempts = int(self.config['SerialConfig'].get('max_attempts', '15'))
        self.excel_filename = "density_data.xlsx"
        self.product_info_list = []
        self.current_product_index = 0
        self.density_values = []
        self.detecting = False
        self.detect_thread = None
        self.auto_mode = False  # 全自动模式标志
        
        # 创建界面组件
        self.create_widgets()
        
        # 初始化时读取Excel文件
        self.load_excel_file()
    
    def create_widgets(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格布局
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)
        
        # 1. 标题区域
        title_frame = ttk.Frame(main_frame)
        title_frame.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        self.title_label = ttk.Label(title_frame, text="密度检测系统", font=(("Segoe UI", 16, "bold")),
                                    foreground=self.mac_colors["text"])
        self.title_label.pack(side=tk.LEFT, padx=5)
        
        # 提示信息标签
        self.prompt_label = ttk.Label(title_frame, text="", font=(("Segoe UI", 12, "bold")), foreground="red")
        self.prompt_label.pack(side=tk.LEFT, padx=20)
        
        self.status_label = ttk.Label(title_frame, text="就绪", font=(("Segoe UI", 10)))
        self.status_label.pack(side=tk.RIGHT, padx=5)
        
        # 2. 产品列表区域
        product_frame = ttk.LabelFrame(main_frame, text="产品型号列表", padding="5")
        product_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # Excel文件路径选择
        file_frame = ttk.Frame(product_frame)
        file_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.excel_path_var = tk.StringVar(value=self.excel_filename)
        self.excel_path_entry = ttk.Entry(file_frame, textvariable=self.excel_path_var, width=60)
        self.excel_path_entry.pack(side=tk.LEFT, padx=5)
        
        self.browse_button = ttk.Button(file_frame, text="浏览", command=self.browse_excel_file)
        self.browse_button.pack(side=tk.LEFT, padx=5)
        
        self.load_button = ttk.Button(file_frame, text="加载", command=self.load_excel_file)
        self.load_button.pack(side=tk.LEFT, padx=5)
        
        # 产品列表
        list_frame = ttk.Frame(product_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 列表控件
        self.product_list = ttk.Treeview(list_frame, columns=("产品型号", "机台号", "来样时间", "班次"), show="headings")
        self.product_list.heading("产品型号", text="产品型号")
        self.product_list.heading("机台号", text="机台号")
        self.product_list.heading("来样时间", text="来样时间")
        self.product_list.heading("班次", text="班次")
        
        # 绑定事件，确保所有行都使用统一的背景色
        self.product_list.bind("<Configure>", self.on_tree_configure)
        self.product_list.bind("<<TreeviewSelect>>", self.on_tree_select)
        
        self.product_list.column("产品型号", width=150)
        self.product_list.column("机台号", width=100)
        self.product_list.column("来样时间", width=150)
        self.product_list.column("班次", width=100)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.product_list.yview)
        self.product_list.configure(yscroll=scrollbar.set)
        
        # 布局
        self.product_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 3. 控制区域
        control_frame = ttk.Frame(main_frame)
        control_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=5)
        
        self.start_button = ttk.Button(control_frame, text="开始检测", command=self.start_detection)
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = ttk.Button(control_frame, text="停止检测", command=self.stop_detection, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        self.next_button = ttk.Button(control_frame, text="下一个产品", command=self.next_product)
        self.next_button.pack(side=tk.LEFT, padx=5)
        
        self.reset_button = ttk.Button(control_frame, text="重置", command=self.reset_detection)
        self.reset_button.pack(side=tk.LEFT, padx=5)
        
        # 全自动模式复选框
        self.auto_mode_var = tk.BooleanVar(value=False)
        self.auto_mode_check = ttk.Checkbutton(control_frame, text="全自动模式", variable=self.auto_mode_var, command=self.toggle_auto_mode)
        # 设置Checkbutton样式
        self.style.configure(
            "TCheckbutton",
            foreground=self.mac_colors["text"],
            font=("Segoe UI", 10)
        )
        self.style.map(
            "TCheckbutton",
            foreground=[("active", self.mac_colors["text"])]
        )
        self.auto_mode_check.pack(side=tk.LEFT, padx=5)
        
        # 串口配置
        serial_frame = ttk.LabelFrame(control_frame, text="串口配置", padding="5")
        serial_frame.pack(side=tk.RIGHT, padx=5)
        
        # 串口名称
        serial_row1 = ttk.Frame(serial_frame)
        serial_row1.pack(fill=tk.X, pady=2)
        ttk.Label(serial_row1, text="串口: ").pack(side=tk.LEFT, padx=5)
        self.serial_port_var = tk.StringVar(value=self.serial_port)
        self.serial_port_entry = ttk.Entry(serial_row1, textvariable=self.serial_port_var, width=10)
        self.serial_port_entry.pack(side=tk.LEFT, padx=5)
        
        # 波特率
        ttk.Label(serial_row1, text="波特率: ").pack(side=tk.LEFT, padx=5)
        self.baudrate_var = tk.IntVar(value=self.baudrate)
        self.baudrate_combo = ttk.Combobox(serial_row1, textvariable=self.baudrate_var, 
                                            values=[9600, 19200, 38400, 57600, 115200], width=8)
        self.baudrate_combo.pack(side=tk.LEFT, padx=5)
        
        # 数据位
        serial_row2 = ttk.Frame(serial_frame)
        serial_row2.pack(fill=tk.X, pady=2)
        ttk.Label(serial_row2, text="数据位: ").pack(side=tk.LEFT, padx=5)
        self.bytesize_var = tk.IntVar(value=self.bytesize)
        self.bytesize_combo = ttk.Combobox(serial_row2, textvariable=self.bytesize_var, 
                                            values=[5, 6, 7, 8], width=5)
        self.bytesize_combo.pack(side=tk.LEFT, padx=5)
        
        # 停止位
        ttk.Label(serial_row2, text="停止位: ").pack(side=tk.LEFT, padx=5)
        self.stopbits_var = tk.DoubleVar(value=self.stopbits)
        self.stopbits_combo = ttk.Combobox(serial_row2, textvariable=self.stopbits_var, 
                                            values=[1, 1.5, 2], width=5)
        self.stopbits_combo.pack(side=tk.LEFT, padx=5)
        
        # 校验位
        ttk.Label(serial_row2, text="校验位: ").pack(side=tk.LEFT, padx=5)
        self.parity_var = tk.StringVar(value=self.parity)
        self.parity_combo = ttk.Combobox(serial_row2, textvariable=self.parity_var, 
                                         values=['NONE', 'ODD', 'EVEN'], 
                                         width=8)
        self.parity_combo.pack(side=tk.LEFT, padx=5)
        
        # 重试次数
        serial_row3 = ttk.Frame(serial_frame)
        serial_row3.pack(fill=tk.X, pady=2)
        ttk.Label(serial_row3, text="重试次数: ").pack(side=tk.LEFT, padx=5)
        self.max_attempts_var = tk.IntVar(value=self.max_attempts)
        self.max_attempts_entry = ttk.Entry(serial_row3, textvariable=self.max_attempts_var, width=10)
        self.max_attempts_entry.pack(side=tk.LEFT, padx=5)
        
        # 保存配置按钮
        save_button = ttk.Button(serial_row3, text="保存配置", command=self.save_config)
        save_button.pack(side=tk.RIGHT, padx=5)
        
        # 4. 数据显示区域
        display_frame = ttk.LabelFrame(main_frame, text="检测数据", padding="5")
        display_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # 左侧：原始数据和测试结果
        left_frame = ttk.Frame(display_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        # 原始数据
        raw_frame = ttk.LabelFrame(left_frame, text="原始数据", padding="5")
        raw_frame.pack(fill=tk.BOTH, expand=False, padx=5, pady=5)  # 设置expand=False，不自动扩展
        
        self.raw_data_text = scrolledtext.ScrolledText(raw_frame, width=60, height=5, font=("Courier New", 10),
                                                      foreground=self.mac_colors["text"],
                                                      insertbackground=self.mac_colors["primary"],
                                                      relief="flat",
                                                      borderwidth=1)
        self.raw_data_text.pack(fill=tk.BOTH, expand=True)
        
        # 检测结果
        result_frame = ttk.LabelFrame(left_frame, text="检测结果", padding="5")
        result_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        result_table_frame = ttk.Frame(result_frame)
        result_table_frame.pack(fill=tk.BOTH, expand=True)
        
        # 检测结果表格
        self.result_table = ttk.Treeview(result_table_frame, columns=("检测次数", "密度值"), show="headings")
        self.result_table.heading("检测次数", text="检测次数")
        self.result_table.heading("密度值", text="密度值 (g/ccm)")
        self.result_table.column("检测次数", width=80, anchor=tk.CENTER)
        self.result_table.column("密度值", width=120, anchor=tk.CENTER)
        
        # 滚动条
        result_scrollbar = ttk.Scrollbar(result_table_frame, orient=tk.VERTICAL, command=self.result_table.yview)
        self.result_table.configure(yscroll=result_scrollbar.set)
        
        # 布局
        self.result_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        result_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 平均值显示
        avg_frame = ttk.Frame(result_frame)
        avg_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(avg_frame, text="平均值: ").pack(side=tk.LEFT, padx=5)
        self.avg_value_var = tk.StringVar(value="--")
        self.avg_value_label = ttk.Label(avg_frame, textvariable=self.avg_value_var, font=(("Segoe UI", 12, "bold")),
                                        foreground=self.mac_colors["primary"])
        self.avg_value_label.pack(side=tk.LEFT, padx=5)
        
        # 右侧：日志和提示信息
        right_frame = ttk.LabelFrame(display_frame, text="操作日志", padding="5")
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5)
        
        self.log_text = scrolledtext.ScrolledText(right_frame, width=60, height=25, font=("Courier New", 10),
                                                foreground=self.mac_colors["text"],
                                                insertbackground=self.mac_colors["primary"],
                                                relief="flat",
                                                borderwidth=1)
        self.log_text.pack(fill=tk.BOTH, expand=True)
    
    def browse_excel_file(self):
        """选择Excel文件路径"""
        filename = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.excel_path_var.set(filename)
            self.excel_filename = filename
            self.load_excel_file()
    
    def load_excel_file(self):
        """加载Excel文件并显示产品列表"""
        self.log_message("正在加载Excel文件...")
        self.status_label.config(text="加载中")
        
        try:
            if not self.excel_filename or not os.path.exists(self.excel_filename):
                for item in self.product_list.get_children():
                    self.product_list.delete(item)
                self.product_info_list = []
                self.current_product_index = 0
                self.log_message("未找到Excel文件，请先选择一个.xlsx文件")
                self.status_label.config(text="就绪")
                return

            self.product_info_list = read_product_models_from_excel(self.excel_filename)
            
            # 清空产品列表
            for item in self.product_list.get_children():
                self.product_list.delete(item)
            
            # 添加产品到列表
            for info in self.product_info_list:
                self.product_list.insert("", tk.END, values=(
                    info["产品型号"],
                    info["机台号"],
                    info["来样时间"],
                    info["班次"]
                ))
            

            
            self.log_message(f"成功加载 {len(self.product_info_list)} 个产品型号")
            self.status_label.config(text="就绪")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载Excel文件失败: {str(e)}")
            self.log_message(f"加载Excel文件失败: {str(e)}")
            self.status_label.config(text="错误")
    
    def start_detection(self):
        """开始检测"""
        if self.detecting:
            return
        
        if not self.product_info_list:
            messagebox.showwarning("警告", "未加载任何产品型号")
            return
        
        if self.current_product_index >= len(self.product_info_list):
            messagebox.showinfo("提示", "所有产品都已检测完成")
            return
        
        # 获取当前产品信息
        current_product = self.product_info_list[self.current_product_index]
        product_model = current_product["产品型号"]
        
        # 在主界面显示提示信息
        self.prompt_label.config(text=f"请放入 {product_model} 型号的样块")
        
        # 更新界面状态
        self.detecting = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.next_button.config(state=tk.DISABLED)
        self.reset_button.config(state=tk.DISABLED)
        
        # 重置检测数据
        self.density_values = []
        self.clear_detection_results()
        
        # 更新日志
        self.log_message(f"开始检测产品: {product_model}")
        self.status_label.config(text="检测中")
        
        # 清空提示信息
        self.root.after(1000, lambda: self.prompt_label.config(text=""))
        
        # 启动检测线程
        self.detect_thread = threading.Thread(target=self.run_detection)
        self.detect_thread.daemon = True
        self.detect_thread.start()
    
    def save_config(self):
        """保存配置到文件"""
        try:
            # 更新配置
            self.config['SerialConfig']['port'] = self.serial_port_var.get()
            self.config['SerialConfig']['baudrate'] = str(self.baudrate_var.get())
            self.config['SerialConfig']['bytesize'] = str(self.bytesize_var.get())
            self.config['SerialConfig']['stopbits'] = str(self.stopbits_var.get())
            self.config['SerialConfig']['parity'] = self.parity_var.get()
            self.config['SerialConfig']['timeout'] = str(self.timeout)
            self.config['SerialConfig']['max_attempts'] = str(self.max_attempts_var.get())
            
            # 保存到文件
            with open(self.config_file, 'w') as f:
                self.config.write(f)
            
            # 更新内存中的配置
            self.max_attempts = self.max_attempts_var.get()
            
            messagebox.showinfo("提示", "串口配置已保存")
            self.log_message("串口配置已保存到文件")
        except Exception as e:
            messagebox.showerror("错误", f"保存配置失败: {e}")
            self.log_message(f"保存配置失败: {e}")
    
    def run_detection(self):
        """执行检测流程"""
        try:
            current_product = self.product_info_list[self.current_product_index]
            product_model = current_product["产品型号"]
            machine_id = current_product["机台号"]
            sample_time = current_product["来样时间"]
            shift = current_product["班次"]
            
            # 更新串口参数
            self.serial_port = self.serial_port_var.get()
            self.baudrate = self.baudrate_var.get()
            self.bytesize = self.bytesize_var.get()
            self.stopbits = self.stopbits_var.get()
            self.parity = self.parity_var.get()
            
            # 开始5次密度检测
            density_values = []
            detect_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for detect_num in range(1, 6):
                if not self.detecting:
                    break
                
                self.log_message(f"开始第 {detect_num} 次检测...")
                
                # 读取串口数据，优化重试机制
                max_attempts = self.max_attempts_var.get()  # 从界面获取重试次数
                density = None
                raw_data = ""
                
                for attempt in range(max_attempts):
                    if not self.detecting:
                        break
                    
                    raw_data = read_serial_data(
                        self.serial_port,
                        baudrate=self.baudrate,
                        bytesize=self.bytesize,
                        stopbits=self.stopbits,
                        parity=self.parity,
                        timeout=3  # 延长单次读取超时时间
                    )
                    if raw_data:
                        # 更新原始数据显示
                        self.root.after(0, self.update_raw_data, raw_data)
                        self.log_message(f"第 {detect_num} 次检测 - 第 {attempt + 1} 次尝试读取到原始数据")
                        
                        # 提取密度值
                        density = extract_density_value(raw_data)
                        if density is not None:
                            self.log_message(f"第 {detect_num} 次检测 - 成功提取密度值: {density} g/ccm")
                            break
                        else:
                            # 即使没有找到密度值，也稍微等待一下
                            self.log_message(f"第 {detect_num} 次检测 - 读取到数据但未找到密度值，重试中...")
                            time.sleep(0.5)
                    else:
                        self.log_message(f"第 {detect_num} 次检测 - 第 {attempt + 1} 次尝试读取失败，重试中...")
                        # 指数退避策略
                        wait_time = min(0.1 * (2 ** attempt), 2)  # 最大等待2秒
                        time.sleep(wait_time)
                
                if density is not None:
                    density_values.append(density)
                    # 更新检测结果表格
                    self.root.after(0, self.add_detection_result, detect_num, density)
                else:
                    density_values.append(None)
                    self.root.after(0, self.add_detection_result, detect_num, "失败")
                    self.log_message(f"第 {detect_num} 次检测 - 失败")
                
                # 获取到数据后不等待，直接进行下一次检测
                # # 如果需要等待，可以调整这里的时间间隔
                # time.sleep(0.1)
            
            # 如果检测完成
            if self.detecting:
                # 计算平均值（仅包含有效数值）
                valid_densities = [d for d in density_values if d is not None]
                average_density = sum(valid_densities) / len(valid_densities) if valid_densities else None
                
                # 更新平均值显示
                avg_str = f"{average_density:.4f}" if average_density is not None else "--"
                self.root.after(0, self.avg_value_var.set, avg_str)
                
                # 准备检测数据
                detect_data = {
                    "来样时间": sample_time,
                    "检测时间": detect_time,
                    "机台号": machine_id,
                    "产品型号": product_model,
                    "班次": shift,
                    "密度1": density_values[0] if len(density_values) > 0 else None,
                    "密度2": density_values[1] if len(density_values) > 1 else None,
                    "密度3": density_values[2] if len(density_values) > 2 else None,
                    "密度4": density_values[3] if len(density_values) > 3 else None,
                    "密度5": density_values[4] if len(density_values) > 4 else None,
                    "平均值": round(average_density, 4) if average_density is not None else None
                }
                
                # 更新Excel文件
                update_excel_with_detection_results(self.excel_filename, product_model, detect_data)
                self.log_message(f"成功更新 {product_model} 的检测结果到Excel文件")
                
                # 更新界面状态
                self.root.after(0, self.detection_completed)
        
        except Exception as e:
            self.root.after(0, messagebox.showerror, "测试错误", f"测试过程中发生错误: {str(e)}")
            self.root.after(0, self.log_message, f"测试错误: {str(e)}")
            self.root.after(0, self.stop_detection)
    
    def stop_detection(self):
        """停止检测"""
        self.detecting = False
        
        # 等待检测线程结束
        if self.detect_thread and self.detect_thread.is_alive():
            self.detect_thread.join(timeout=1.0)
        
        # 更新界面状态
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.next_button.config(state=tk.NORMAL)
        self.reset_button.config(state=tk.NORMAL)
        
        self.status_label.config(text="已停止")
        self.log_message("检测已停止")
    
    def next_product(self):
        """检测下一个产品"""
        if self.detecting:
            messagebox.showwarning("警告", "当前正在检测，请先停止检测")
            return
        
        if self.current_product_index < len(self.product_info_list) - 1:
            self.current_product_index += 1
            self.clear_detection_results()
            self.log_message(f"切换到第 {self.current_product_index + 1} 个产品")
        else:
            messagebox.showinfo("提示", "已经是最后一个产品")
    
    def reset_detection(self):
        """重置检测状态"""
        if self.detecting:
            self.stop_detection()
        
        self.current_product_index = 0
        self.density_values = []
        self.clear_detection_results()
        self.log_message("检测已重置")
    
    def toggle_auto_mode(self):
        """切换全自动模式"""
        self.auto_mode = self.auto_mode_var.get()
        self.log_message(f"{'启用' if self.auto_mode else '禁用'}全自动模式")
        
    def on_tree_configure(self, event):
        """Treeview配置变化时的处理"""
        # 尝试更新Treeview的背景色
        try:
            # 获取Treeview的内部组件
            tree_widget = self.product_list
            # 强制刷新
            tree_widget.update_idletasks()
        except Exception as e:
            pass
        
    def on_tree_select(self, event):
        """Treeview选择变化时的处理"""
        # 当选择变化时，确保选中行的样式正确
        pass
    
    def detection_completed(self):
        """
        检测完成后的处理
        """
        self.detecting = False
        
        # 清空提示信息
        self.prompt_label.config(text="")
        
        # 更新界面状态
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.next_button.config(state=tk.NORMAL)
        self.reset_button.config(state=tk.NORMAL)
        
        current_product = self.product_info_list[self.current_product_index]
        product_model = current_product["产品型号"]
        
        if self.auto_mode:
            # 如果启用了全自动模式，自动检测下一个产品
            self.status_label.config(text="检测完成，准备下一个产品")
            self.log_message(f"{product_model} 型号检测完成，准备下一个产品...")
            
            # 等待1秒，然后自动开始下一个产品的检测
            if self.current_product_index < len(self.product_info_list) - 1:
                self.root.after(1000, self.auto_next_product)
            else:
                # 所有产品检测完成
                self.status_label.config(text="所有产品检测完成")
                messagebox.showinfo("检测完成", "所有产品的检测已完成")
                self.log_message("所有产品检测完成")
        else:
            # 非全自动模式，正常显示检测完成信息
            self.status_label.config(text="检测完成")
            messagebox.showinfo("检测完成", f"{product_model} 型号的检测已完成")
    
    def auto_next_product(self):
        """全自动模式下自动开始下一个产品的检测"""
        if self.detecting:
            return
        
        if self.current_product_index < len(self.product_info_list) - 1:
            self.current_product_index += 1
            self.clear_detection_results()
            self.log_message(f"自动切换到第 {self.current_product_index + 1} 个产品")
            
            # 自动开始检测
            self.start_detection()
        else:
            # 所有产品检测完成
            self.status_label.config(text="所有产品检测完成")
            messagebox.showinfo("检测完成", "所有产品的检测已完成")
            self.log_message("所有产品检测完成")
    
    def update_raw_data(self, data):
        """更新原始数据显示"""
        self.raw_data_text.delete("1.0", tk.END)
        self.raw_data_text.insert(tk.END, data)
    
    def add_detection_result(self, detect_num, value):
        """添加检测结果到表格"""
        self.result_table.insert("", tk.END, values=(f"第 {detect_num} 次", value))
    
    def clear_detection_results(self):
        """清空检测结果"""
        # 清空原始数据
        self.raw_data_text.delete("1.0", tk.END)
        
        # 清空检测结果表格
        for item in self.result_table.get_children():
            self.result_table.delete(item)
        
        # 清空平均值
        self.avg_value_var.set("--")
    
    def log_message(self, message):
        """添加日志信息"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)


# 主函数调用
if __name__ == "__main__":
    # 方式1：实际串口读取（注释掉方式2和3，启用此行）
    # main()

    # 方式2：使用固定数据测试（注释掉方式1和3，启用此行）
    # test_with_fixed_data()
    
    # 方式3：运行GUI界面（注释掉方式1和2，启用此行）
    try:
        root = tk.Tk()
        app = DensityDetectGUI(root)
        root.mainloop()
    except Exception as e:
        print(f"GUI应用运行出错: {e}")
        import traceback
        traceback.print_exc()
