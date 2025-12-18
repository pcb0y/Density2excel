import openpyxl

# 加载Excel文件
wb = openpyxl.load_workbook('density_data.xlsx')
ws = wb.active

# 打印表头
print('表头:')
headers = []
for cell in ws[1]:
    headers.append(cell.value)
print(headers)

# 打印所有数据行
print('\n所有数据行:')
for row in ws.iter_rows(min_row=2, values_only=True):
    print(row)

# 关闭工作簿
wb.close()